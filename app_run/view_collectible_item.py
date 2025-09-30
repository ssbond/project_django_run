from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import CollectibleItem
from .serializers import CollectibleItemSerializer
from rest_framework.decorators import api_view
from openpyxl import load_workbook

class CollectibleItemApiViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer
    pass

class UploadFileApiView(APIView):
    pass

@api_view(['POST'])
def upload_collectible_items_xls(request):
    file = request.FILES.get('file')
    if not file:
        return Response({"error": "No file provided."}, status=400)

    workbook = load_workbook(file)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)  # Получаем заголовки столбцов
    header_data = []
    for h in header:
        header_data.append(h.lower())
        if h=='URL':
            header_data[-1]='picture'
    filedata = []
    for row in rows:
        filedata.append(dict(zip(header_data, row)))
    invalid_rows = []  # Здесь будем хранить данные невалидных строк
    for i, row in enumerate(filedata, start=2):  # start=2, т.к. 1 строка - заголовок
        serializer = CollectibleItemSerializer(data=row)
        if serializer.is_valid():
            serializer.save()
        else:
            original_row_data = [row[field_name] for field_name in header_data]
            print(row)
            # print(field_name)
            print(header_data)


            invalid_rows.append(original_row_data)
    return Response(invalid_rows, status=200)
